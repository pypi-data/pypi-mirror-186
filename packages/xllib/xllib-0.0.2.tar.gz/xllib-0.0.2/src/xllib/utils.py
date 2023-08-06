import csv
import json
import openpyxl
from copy import deepcopy
from typing import Sequence

def get_tbl_data(tbl: list):
    """Extract list data from class CT if given
    """
    if type(tbl).__name__ == 'CT':
        return tbl.x
    if isinstance(tbl, list):
        return tbl
    raise Exception("Unexpected data structure!")

def tbl_rbind(tbls: list, fill=""):
    max_row_len = max(len(row) for tbl in tbls for row in get_tbl_data(tbl))
    merged = []
    for tbl in deepcopy(tbls):
        for row in get_tbl_data(tbl):
            lr = len(row)
            if lr < max_row_len:
                row += [fill] * (max_row_len - lr)
            merged.append(row)
    return merged

def transpose(tbl):
    return [ list(x) for x in zip(*get_tbl_data(tbl)) ]

def tbl_cbind(tbls: list, fill=""):
    tbls_tr = [ transpose(get_tbl_data(x)) for x in tbls ]
    return transpose(tbl_rbind(tbls_tr, fill=fill))

def write_csv(x, fp, encoding="utf-8", sep=","):
    with open(fp, "w", encoding=encoding, newline='') as f:
        writer = csv.writer(f, delimiter=sep)
        writer.writerows(get_tbl_data(x))

def read_csv(fp, encoding="utf-8", delimiter=","):
    with open(fp, encoding=encoding, newline='') as f:
        reader = csv.reader(f, delimiter=delimiter)
        return [ r for r in reader ]

def read_json_tbl(fp, columns:list=None):
    '''
    [
        {A: 1, B: 2, C: 3},
        {A: 4, B: 5, C: 6},
        ...
    ]
    '''
    with open(fp, encoding="utf-8") as f:
        d = json.load(f)
    if columns is None: columns = list(d[0].keys())
    tbl = []
    for row in d:
        new_row = [ row[col] for col in columns ]
        tbl.append(new_row)
    return tbl


def readxl(fp, sheet=0):
    wb = openpyxl.load_workbook(fp)
    if isinstance(sheet, int):
        ws = wb.sheetnames[sheet]
        ws = wb[ws]
    elif isinstance(sheet, str):
        ws = wb[sheet]
    else:
        raise TypeError("`sheet` should be int or str!")
    tbl = []
    for row in ws.iter_rows():
        row_lst = [ x.value for x in row ]
        tbl.append(row_lst)
    return tbl


def writexls(tbls: list, fp, sheetnames=None):
    if isinstance(sheetnames, str):
        sheetnames = [ f"{sheetnames}_{i+1}" for i in range(len(tbls)) ]
    if isinstance(sheetnames, Sequence):
        if len(sheetnames) != len(tbls):
            raise Exception("Different lengths of given sheetnames and tables!")
    if sheetnames is None: sheetnames = [ f"Sheet{i+1}" for i in range(len(tbls)) ]
    
    wb = openpyxl.Workbook()
    ws0 = wb.active
    wb.remove(ws0)
    for i, tbl in enumerate(tbls):
        ws = wb.create_sheet(sheetnames[i])
        for row in get_tbl_data(tbl):
            ws.append(row)
    wb.save(fp)


def writexl(tbl: list, fp, sheetname:str=None):
    if sheetname is None: sheetname = "Sheet"
    writexls(tbls=[tbl], fp=fp, sheetnames=[sheetname])
