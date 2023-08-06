from openpyxl import load_workbook
import sys
import xlrd
from openpyxl.workbook import Workbook as openpyxlWorkbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE


def convierte_xls_to_xlsx(content):
    xlsBook = xlrd.open_workbook(content)
    workbook = openpyxlWorkbook()

    for i in range(0, xlsBook.nsheets):
        xlsSheet = xlsBook.sheet_by_index(i)
        sheet = workbook.active if i == 0 else workbook.create_sheet()
        sheet.title = xlsSheet.name

        for row in range(0, xlsSheet.nrows):
            for col in range(0, xlsSheet.ncols):
                sheet.cell(row=row+1, column=col+1).value = ILLEGAL_CHARACTERS_RE.sub(r' ', str(xlsSheet.cell_value(row, col)))

    workbook.save(content+"x")

def unifica_hojas(archivo, desplazamiento):
    print("Conviertiendo de xls a xlsx")
    convierte_xls_to_xlsx(archivo)

    print("Unificando hojas en una sóla hoja")
    book = load_workbook(filename=archivo+"x")
    hojas = book.sheetnames
    print(hojas)
    destino = book[hojas[0]]
    if len(hojas) > 1:   # Hay mas de una hoja
        for org in hojas[1:]:
            origen = book[org]
            print("De ", org, " a ", hojas[0], " en ", destino.max_row + 1)
            fila_destino = destino.max_row
            for fila in origen.iter_rows(min_row = desplazamiento):
                fila_destino += 1
                for cell in fila:
                    destino.cell(row = fila_destino, column = cell.column).value = cell.value
            book.remove(origen)

    if desplazamiento == 4:  # Hay cabecera de búsqueda
        print("Eliminando cabecera de búsqueda")
        destino.delete_rows(1, amount=2)
    book.save(archivo+"x")



'''
if len(sys.argv) != 3:
    print("Error: usage Convert_XLS_to_XLSX file.xls eliminar_busqueda|no_eliminar_busqueda")
    unifica_hojas("c:\\local\\Docuconta_2022.xls", 4)
else:
    fichero = sys.arv[1]
    busqueda = sys.argv[2]
    if busqueda == "eliminar_busqueda":
        unifica_hojas(fichero, 4)
    else:
        unifica_hojas(fichero, 2)

'''