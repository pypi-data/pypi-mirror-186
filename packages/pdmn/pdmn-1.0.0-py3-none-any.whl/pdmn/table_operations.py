"""
This file is part of the pDMN solver.
Author: Simon Vandevelde
s.vandevelde@kuleuven.be
"""
from typing import List, Dict
import openpyxl
import itertools
import numpy as np
from pdmn.table import Table
from pdmn.glossary import Glossary
from pdmn.idply import Parser
import re


def fill_in_merged(file_name: str, sheet_names: List[str] = None) \
                                                -> List[np.array]:
    """
    Loads up a specific sheet and returns it.
    A sheet is comprised of a list of tables.

    :arg file_name: path to the xlsx file.
    :arg sheet: name of the sheet to load in.
    :returns List<np.array>: a list of all the tables in a sheet.
    """
    def rang(m, n): return range(m, n + 1)

    wb = openpyxl.load_workbook(file_name)
    sheets = []
    for sheet_name in sheet_names:
        sheets.append(wb[sheet_name])

    for sheet in sheets:
        ranges = list(sheet.merged_cells.ranges)
        for xyrange in ranges:
            b = xyrange.bounds
            sheet.unmerge_cells(str(xyrange))

            for x, y in itertools.product(rang(b[0], b[2]), rang(b[1], b[3])):
                sheet.cell(y, x).value = sheet.cell(xyrange.min_row,
                                                    xyrange.min_col).value
    return [np.array(list(s.values)) for s in sheets]


def find_first(sheet: np.array) -> tuple:
    """
    TODO

    :arg sheet: the sheet
    """
    for x, y in itertools.product(*[list(range(s)) for s in sheet.shape]):
        if sheet[x, y]:
            return x, y
    return (None, None)


def explore(sheet: List[np.array], boundaries: List[int]):
    """
    Tries to find the ranges for each table.

    :arg sheet: the sheet containing all the tables.
    :arg bounaries: a list containing the theoretic boundaries.
    :returns None:
    """
    startx = boundaries[0]
    endx = boundaries[2]
    starty = boundaries[1]
    endy = boundaries[3]

    # This is a hack to make sure data tables with no input can get found.
    if re.match(r"(?i)DataTable|Data Table", sheet[startx, starty],
                re.IGNORECASE):
        endx += 1
        endy += 1

    while True:
        changed = False

        try:
            while any(sheet[endx+1, starty: endy+1]):
                endx += 1
                changed = True
        except IndexError:
            pass
        try:
            while any(sheet[startx: endx+1, endy+1]):
                endy += 1
                changed = True
        except IndexError:
            pass
        if not changed:
            break
    boundaries[2] = endx + 1
    boundaries[3] = endy + 1


def identify_tables(sheets: List[np.array]) -> List[np.array]:
    """
    Function which looks for all the tables in a given sheet.
    Creates a list of boundaries for the tables.

    :arg sheets: a list containing a numpy array representing the sheet.
    :returns List[np.array]: containing boundaries of all the tables.
    """
    tables = []
    for sheet in sheets:
        while True:
            index = find_first(sheet)
            if index[0] is None and index[1] is None:
                break
            boundaries = list(index + index)
            explore(sheet, boundaries)
            tables.append(sheet[boundaries[0]: boundaries[2],
                          boundaries[1]: boundaries[3]].copy())
            sheet[boundaries[0]: boundaries[2],
                  boundaries[1]: boundaries[3]] = None

    return tables


def find_tables_by_name(tables, name):
    """
    Looks for tables based on a regex expression representing their name.

    :arg np.array: the tables.
    :arg str: the name, in regex.
    :returns List<np.array>: the tables found.
    """
    named_tables = []
    for table in tables:
        if re.match(name, table[0, 0], re.IGNORECASE):
            named_tables.append(table)
    return named_tables


def find_glossary(tables: List[np.array]) -> Dict[str, np.array]:
    """
    Locates the glossarytables, and places them in a dictionary for each type.
    The three tables it looks for are:

        * Type
        * Function
        * Predicate

    :returns Dict[str, np.array]: containing the glossary for type, function,
    constant, predicate and boolean.
    """
    glossary = {"Type": None, "Function": None, "Predicate": None}

    glossary["Type"] = find_glossary_table(tables, "type", critical=True)
    glossary["Function"] = find_glossary_table(tables, "function")
    glossary["Predicate"] = find_glossary_table(tables, "predicate")

    return glossary


def find_glossary_table(tables: List[np.array], name: str,
                        critical: bool = False) -> np.array:
    """
    Looks for a specific glossary table with name "name".

    If the critical boolean is set, an error is returned when no glossary is
    found. For example, there should always be a type glossary.
    Non-critical glossaries only print warnings when none are found.

    :arg tables: the list of arrays, each containing a table.
    :arg name: the name of the table to find.
    :arg critical: True if a table needs to be found. E.g. if the table is not
        found, an error is thrown.
    :returns np.array: the glossary table, if found.
    """
    glossaries = find_tables_by_name(tables, name)
    if len(glossaries) == 0:
        if critical:
            raise ValueError(f"No {name} glossary table was found.")
        else:
            print(f"INFO: No {name} glossary table found.")
        return None

    if len(glossaries) > 1:
        raise ValueError(f"Multiple {name} glossary tables were found.")
    return glossaries[0]


def find_execute_method(tables: List[np.array]) -> Dict[str, object]:
    """
    Locates the Query table, which contains the predicates for which we want to
    know their probabilities.

    :arg tables: the list of arrays, each containing a table.
    :returns List[str]: contains the target predicates
    """
    query_table = find_tables_by_name(tables, 'Query')
    if len(query_table) > 1:
        raise ValueError("Only one Query table allowed")
    # Grab al predicates listed in the query table.
    if len(query_table) == 0:
        return []
    query_preds = [x[0] for x in query_table[0][1:]]
    return query_preds


def find_tables(tables: List[np.array]) -> List[np.array]:
    """
    Looks for decision tables and constraint tables.

    :arg tables: the list of arrays, each containing a table.
    :returns List[np.array]: a list containing only the decision and constraint
        tables.
    """
    tables = list(filter(lambda table:
                         not re.match('Glossary|DataTable|Comment|Data Table',
                                      table[0, 0],
                                      re.IGNORECASE),
                         tables))
    return tables


def create_theory(tables: List[np.array], parser: Parser,
                  glossary: Glossary, query_preds=[]) -> str:
    """
    Function to create the theory in the ProbLog format.

    :arg tables: the list of arrays, each containing a table.
    :arg parser: the parser.
    :returns str: the theory for the ProbLog file.
    """
    theory = "% facts\n" + glossary.to_theory() + "\n"

    inner_t = ""
    for dt in find_tables(tables):
        t = Table(dt, parser).export()
        if t is None:
            continue
        inner_t += t

    theory += inner_t

    # If a query table is present, turn it into query statements.
    if query_preds:
        for query_pred in query_preds:
            pred = parser.parse_val(query_pred, 'Yes')
            theory += f"query({pred}).\n"
    else:
        # if no query is present, just query everything.
        for pred in glossary.predicates:
            pred = parser.parse_val(pred.name, 'Yes')
            theory += f"query({pred}).\n"

    return theory


def create_dependency_graph(tables: List[np.array], parser: Parser) -> Dict:
    """
    Function to create dependency graph of variables.
    Basically, we check for every output of a decision table what the inputs
    are. The outputs depend on these inputs.
    By doing this for every table, we can build a graph of what knowledge
    depends on what other knowledge.

    :arg tables: the list of arrays, each containing a table.
    :arg parser: the parser.
    :returns str: the theory for the ProbLog file.
    """

    dependency_graph: Dict = {}
    for dt in find_tables(tables):
        t = Table(dt, parser, False)

        # Only decision tables count as dependency.
        if t.hit_policy == "E*":
            continue
        for outp in t.outputs:
            if outp not in dependency_graph:
                dependency_graph[outp] = t.inputs
            else:
                for inp in t.inputs:
                    dependency_graph[outp].append(inp)
        if t is None:
            continue

    return dependency_graph
