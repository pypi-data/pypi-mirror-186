"""
excel
=====

Quick description
"""


from openpyxl import load_workbook
import re
import pandas as pd


class ExcelFile():

    def __init__(self, filepath: str) -> None:
        self.filepath = filepath
        self.workbook = load_workbook(filename=filepath, data_only=True, keep_links=False)

    def __repr__(self) -> str:
        myString = "Instance de la classe ExcelFile\n- Nom du fichier : {}".format(self.filepath)
        return myString

    def get_sheet(self):
        return self.workbook.sheetnames

    def get_table(self, sheet_name, table_name):
        a = self.workbook[sheet_name].tables[table_name].ref.split(':')

        for i in range(len(a)):
            a[i] = re.split("(\\d+)", a[i])[:-1]
        df = pd.read_excel(self.filepath, sheet_name=sheet_name, skiprows=int(a[0][1]) - 1, nrows=int(a[1][1])-int(a[0][1]),
                           usecols=a[0][0] + ":" + a[1][0])
        return df

    def get_value(self, value_name):
        temp = self.workbook.defined_names[value_name].attr_text.split('$', 1)
        sheet = temp[0][:-1]
        cell = temp[1]
        return self.workbook[sheet][cell].value

    def get_range(self, sheet_name, cols, skiprows, nrows):
        """
        Parameters
        ----------
        - sheet_name (str) \t Name of the sheet
        - cols (str) \t\t Columns used (ex : 'A:N' or 'A, C, F:T")
        - skiprows (int) \t Number of rows to skip
        - nrows (int) \t\t Length of the Excel tab

        Return
        ------
        * (df) \t Excel range in a pandas dataframe.

        Example
        -------
        >>> objet.get_range(sheet_name="My Sheet", cols='F:M', skiprows=5, nrows=10)

        """
        return pd.read_excel(self.filepath, sheet_name=sheet_name, usecols=cols, skiprows=skiprows, nrows=nrows)

    def write_cell1(self, sheet_name, cell, value) -> None:
        self.workbook[sheet_name][cell].value = value

    def write_cell2(self, sheet_name, row, col, value) -> None:
        self.workbook[sheet_name].cell(row=row, column=col).value = value

    def save(self) -> None:
        self.workbook.save(self.filepath)
